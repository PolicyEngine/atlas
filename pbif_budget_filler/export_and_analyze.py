#!/usr/bin/env python3
"""Export the spreadsheet to xlsx and analyze its structure"""

import pickle
import gspread
from pathlib import Path
import requests

def export_and_analyze():
    """Export spreadsheet to xlsx for analysis"""
    
    # Load credentials
    token_path = Path(__file__).parent.parent / "token.pickle"
    with open(token_path, 'rb') as token:
        creds = pickle.load(token)
    
    gc = gspread.authorize(creds)
    
    SPREADSHEET_ID = "1sJdmn3IF09h0YA7hYeem80CCfDc1z8jYdeCkq5Phknw"
    
    print("EXPORTING SPREADSHEET TO XLSX")
    print("="*70)
    
    # Export URL for xlsx
    export_url = f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}/export?format=xlsx"
    
    print(f"\nExport URL: {export_url}")
    
    # Get the file
    headers = {'Authorization': f'Bearer {creds.token}'}
    response = requests.get(export_url, headers=headers)
    
    if response.status_code == 200:
        # Save the xlsx file
        output_file = "pbif_budget_spreadsheet.xlsx"
        with open(output_file, 'wb') as f:
            f.write(response.content)
        
        print(f"✓ Saved to {output_file}")
        
        # Now analyze with openpyxl
        try:
            import openpyxl
        except ImportError:
            print("\nInstalling openpyxl to analyze...")
            import subprocess
            subprocess.check_call(["/opt/homebrew/bin/python3.10", "-m", "pip", "install", "openpyxl"])
            import openpyxl
        
        print("\n" + "="*70)
        print("ANALYZING XLSX STRUCTURE")
        print("="*70)
        
        # Load the workbook
        wb = openpyxl.load_workbook(output_file, data_only=False)  # Keep formulas
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n\nSHEET: {sheet_name}")
            print("-"*40)
            
            # Find cells with formulas
            formulas = []
            data_cells = []
            
            for row in ws.iter_rows(min_row=1, max_row=30, min_col=1, max_col=10):
                for cell in row:
                    if cell.value:
                        if isinstance(cell.value, str) and cell.value.startswith('='):
                            formulas.append(f"  {cell.coordinate}: {cell.value}")
                        elif cell.row > 3:  # Data rows
                            # Check if it's a meaningful data cell
                            if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and len(cell.value) > 2):
                                data_cells.append(f"  {cell.coordinate}: {cell.value}")
            
            if formulas:
                print("\nFORMULAS:")
                for f in formulas[:10]:  # Show first 10
                    print(f)
            
            if data_cells:
                print("\nDATA CELLS:")
                for d in data_cells[:15]:  # Show first 15
                    print(d)
        
        print("\n" + "="*70)
        print("KEY FINDINGS:")
        print("-"*40)
        print("Now we can see the exact structure and formulas!")
        
    else:
        print(f"Error downloading: {response.status_code}")

if __name__ == "__main__":
    export_and_analyze()